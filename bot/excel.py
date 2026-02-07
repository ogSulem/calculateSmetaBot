from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


def build_estimate_xlsx(
    *,
    path: Path,
    area: float,
    items: list[dict[str, Any]],
    total: float,
    price_per_m2: float,
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "Смета"

    headers = ["Раздел", "Описание", "Площадь", "Цена за м²", "Стоимость"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for it in items:
        it_area = float(it.get("area", 0))
        price = float(it.get("price", 0))
        cost = it_area * price
        ws.append(
            [
                str(it.get("section", "")),
                str(it.get("title", it.get("id", ""))),
                it_area,
                price,
                cost,
            ]
        )
        row += 1

    row += 1
    ws.cell(row=row, column=4, value="Итого:").font = header_font
    ws.cell(row=row, column=5, value=total).font = header_font

    row += 1
    ws.cell(row=row, column=4, value="Цена за м²:").font = header_font
    ws.cell(row=row, column=5, value=price_per_m2).font = header_font

    row += 1
    ws.cell(row=row, column=4, value="Площадь:").font = header_font
    ws.cell(row=row, column=5, value=area).font = header_font

    row += 1
    ws.cell(row=row, column=4, value="Дата расчёта:").font = header_font
    ws.cell(row=row, column=5, value=dt.datetime.now().strftime("%Y-%m-%d %H:%M"))

    widths = [16, 38, 12, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord('A') + i - 1)].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
